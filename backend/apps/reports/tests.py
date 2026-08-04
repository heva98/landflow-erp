from datetime import timedelta
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from apps.accounts.models import Role, User
from apps.crm.models import Customer, Lead
from apps.installments.models import PaymentPlan
from apps.plots.models import Plot
from apps.projects.models import Project
from apps.sales.models import Sale

from .views import EXCEL_CONTENT_TYPE, PDF_CONTENT_TYPE


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def sales_manager_role(db):
    return Role.objects.get(name='Sales Manager')


@pytest.fixture
def customer_role(db):
    # Not granted any report-backing permissions in any module's seeding migration.
    return Role.objects.get(name='Customer')


@pytest.fixture
def sales_manager(db, sales_manager_role):
    return User.objects.create_user(email='salesmgr@landflow.co.tz', password='s3cure-pass', role=sales_manager_role)


@pytest.fixture
def customer_user(db, customer_role):
    return User.objects.create_user(email='customer@landflow.co.tz', password='s3cure-pass', role=customer_role)


@pytest.fixture
def project(db):
    return Project.objects.create(
        name='Buyuni Phase II', location='Buyuni, Dar es Salaam', total_area_sqm=Decimal('50000.00'),
    )


@pytest.fixture
def plot(db, project):
    return Plot.objects.create(
        project=project, plot_number='A-01', area_sqm=Decimal('500.00'), price=Decimal('20000000.00'),
        status=Plot.Status.SOLD,
    )


@pytest.fixture
def customer(db):
    return Customer.objects.create(full_name='Juma Hassan', phone='+255700000000')


@pytest.fixture
def sale(db, plot, customer):
    return Sale.objects.create(
        plot=plot, customer=customer, sale_type=Sale.SaleType.CASH,
        sale_price=Decimal('20000000.00'), down_payment=Decimal('20000000.00'),
    )


@pytest.fixture
def payment_plan(db, plot, customer, project):
    installment_plot = Plot.objects.create(
        project=project, plot_number='A-02', area_sqm=Decimal('500.00'), price=Decimal('12000000.00'),
        status=Plot.Status.SOLD,
    )
    installment_sale = Sale.objects.create(
        plot=installment_plot, customer=customer, sale_type=Sale.SaleType.INSTALLMENT,
        sale_price=Decimal('12000000.00'), down_payment=Decimal('2000000.00'),
    )
    plan = PaymentPlan.objects.create(
        sale=installment_sale, principal_amount=Decimal('10000000.00'), interest_rate=Decimal('0'),
        number_of_installments=4, installment_amount=Decimal('2500000.00'),
        start_date=timezone.localdate() + timedelta(days=30),
    )
    plan.generate_schedule()
    return plan


@pytest.fixture
def lead(db, project):
    return Lead.objects.create(
        full_name='Amina Rashid', phone='+255711111111', source=Lead.Source.WEBSITE,
        status=Lead.Status.NEW, interested_project=project,
    )


# --- sales report -------------------------------------------------------------

@pytest.mark.django_db
def test_sales_report_json(api_client, sales_manager, sale):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-sales'))
    assert response.status_code == status.HTTP_200_OK, response.data
    assert response.data['summary']['count'] == 1
    assert Decimal(response.data['summary']['total_sale_price']) == Decimal('20000000.00')
    assert response.data['rows'][0]['sale_number'] == sale.sale_number


@pytest.mark.django_db
def test_sales_report_filters_by_project(api_client, sales_manager, sale):
    other_project = Project.objects.create(
        name='Kigamboni Heights', location='Kigamboni, Dar es Salaam', total_area_sqm=Decimal('10000.00'),
    )
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-sales'), {'project': str(other_project.id)})
    assert response.status_code == status.HTTP_200_OK
    assert response.data['summary']['count'] == 0


@pytest.mark.django_db
def test_sales_report_requires_permission(api_client, customer_user, sale):
    api_client.force_authenticate(user=customer_user)
    response = api_client.get(reverse('report-sales'))
    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
def test_sales_report_requires_authentication(api_client):
    response = api_client.get(reverse('report-sales'))
    assert response.status_code == status.HTTP_401_UNAUTHORIZED


@pytest.mark.django_db
def test_sales_report_excel_export(api_client, sales_manager, sale):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-sales'), {'export': 'xlsx'})
    assert response.status_code == status.HTTP_200_OK
    assert response['Content-Type'] == EXCEL_CONTENT_TYPE
    assert response['Content-Disposition'] == 'attachment; filename="sales-report.xlsx"'

    workbook = openpyxl.load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == 'Sale #'
    assert sheet.cell(row=2, column=1).value == sale.sale_number


@pytest.mark.django_db
def test_sales_report_pdf_export(api_client, sales_manager, sale):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-sales'), {'export': 'pdf'})
    assert response.status_code == status.HTTP_200_OK
    assert response['Content-Type'] == PDF_CONTENT_TYPE
    assert response.content.startswith(b'%PDF')


@pytest.mark.django_db
def test_report_export_rejects_invalid_format(api_client, sales_manager, sale):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-sales'), {'export': 'csv'})
    assert response.status_code == status.HTTP_400_BAD_REQUEST


# --- installment report --------------------------------------------------------

@pytest.mark.django_db
def test_installment_report_json(api_client, sales_manager, payment_plan):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-installments'))
    assert response.status_code == status.HTTP_200_OK, response.data
    assert response.data['summary']['count'] == 4
    assert Decimal(response.data['summary']['total_amount_due']) == payment_plan.total_payable


@pytest.mark.django_db
def test_installment_report_filters_by_status(api_client, sales_manager, payment_plan):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-installments'), {'status': 'paid'})
    assert response.status_code == status.HTTP_200_OK
    assert response.data['summary']['count'] == 0


@pytest.mark.django_db
def test_installment_report_excel_export(api_client, sales_manager, payment_plan):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-installments'), {'export': 'xlsx'})
    assert response.status_code == status.HTTP_200_OK
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    sheet = workbook.active
    # header + 4 installment rows + blank separator + 4 summary lines
    assert sheet.cell(row=2, column=1).value == payment_plan.sale.sale_number


# --- lead report ----------------------------------------------------------------

@pytest.mark.django_db
def test_lead_report_json(api_client, sales_manager, lead):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-leads'))
    assert response.status_code == status.HTTP_200_OK, response.data
    assert response.data['summary']['count'] == 1
    assert response.data['summary']['by_status'] == {'new': 1}
    assert response.data['rows'][0]['full_name'] == 'Amina Rashid'


@pytest.mark.django_db
def test_lead_report_filters_by_source(api_client, sales_manager, lead):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-leads'), {'source': 'referral'})
    assert response.status_code == status.HTTP_200_OK
    assert response.data['summary']['count'] == 0


@pytest.mark.django_db
def test_lead_report_pdf_export(api_client, sales_manager, lead):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-leads'), {'export': 'pdf'})
    assert response.status_code == status.HTTP_200_OK
    assert response.content.startswith(b'%PDF')


# --- land inventory report -------------------------------------------------------

@pytest.mark.django_db
def test_land_inventory_report_json(api_client, sales_manager, plot):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-land-inventory'))
    assert response.status_code == status.HTTP_200_OK, response.data
    assert response.data['summary']['count'] == 1
    assert response.data['summary']['by_status'] == {'sold': 1}
    assert Decimal(response.data['summary']['total_area_sqm']) == Decimal('500.00')


@pytest.mark.django_db
def test_land_inventory_report_filters_by_status(api_client, sales_manager, plot):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-land-inventory'), {'status': 'available'})
    assert response.status_code == status.HTTP_200_OK
    assert response.data['summary']['count'] == 0


@pytest.mark.django_db
def test_land_inventory_report_excel_export(api_client, sales_manager, plot):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-land-inventory'), {'export': 'xlsx'})
    assert response.status_code == status.HTTP_200_OK
    assert response['Content-Disposition'] == 'attachment; filename="land-inventory-report.xlsx"'


# --- cash flow report (delegates to apps.finance.reports.cash_flow) --------------

@pytest.mark.django_db
def test_cash_flow_report_json(api_client, sales_manager):
    api_client.force_authenticate(user=sales_manager)
    today = timezone.localdate()
    response = api_client.get(reverse('report-cash-flow'), {
        'date_from': today.isoformat(), 'date_to': today.isoformat(),
    })
    assert response.status_code == status.HTTP_200_OK, response.data
    assert 'total_inflow' in response.data['summary']


@pytest.mark.django_db
def test_cash_flow_report_requires_date_range(api_client, sales_manager):
    api_client.force_authenticate(user=sales_manager)
    response = api_client.get(reverse('report-cash-flow'))
    assert response.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
def test_cash_flow_report_requires_permission(api_client, customer_user):
    api_client.force_authenticate(user=customer_user)
    today = timezone.localdate()
    response = api_client.get(reverse('report-cash-flow'), {
        'date_from': today.isoformat(), 'date_to': today.isoformat(),
    })
    assert response.status_code == status.HTTP_403_FORBIDDEN
