"""
Excel (openpyxl) and PDF (reportlab) rendering shared by every report. Each
report supplies its column spec (ordered (key, label) pairs matching the row
dicts from reports.py) and a list of (label, value) summary lines; this
module only knows how to lay those out as a spreadsheet or a table.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BRAND_GREEN = colors.HexColor('#0E5B45')
ROW_ALT = colors.HexColor('#F8FAFC')


def _format_cell(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return value


def rows_to_table(rows, columns):
    return [[_format_cell(row.get(key)) for key, _label in columns] for row in rows]


def summary_lines(summary, labels):
    lines = []
    for key, label in labels:
        value = summary.get(key)
        if isinstance(value, dict):
            value = ', '.join(f'{k}: {v}' for k, v in value.items()) or '—'
        elif isinstance(value, Decimal):
            value = f'{value:,.2f}'
        lines.append((label, value))
    return lines


def build_xlsx(title, columns, rows, summary=None):
    headers = [label for _key, label in columns]
    table = rows_to_table(rows, columns)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]

    sheet.append(headers)
    header_fill = PatternFill(start_color='0E5B45', end_color='0E5B45', fill_type='solid')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for row in table:
        sheet.append(row)

    if summary:
        sheet.append([])
        for label, value in summary:
            summary_row = sheet.max_row + 1
            sheet.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=summary_row, column=2, value=value)

    widths = [len(header) for header in headers]
    for row in table:
        for index, value in enumerate(row):
            widths[index] = max(widths[index], len(str(value)) if value is not None else 0)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pdf(title, columns, rows, summary=None):
    headers = [label for _key, label in columns]
    table_data = [headers] + rows_to_table(rows, columns)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=12 * mm, rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 8)]

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    if summary:
        elements.append(Spacer(1, 14))
        for label, value in summary:
            elements.append(Paragraph(f'<b>{label}:</b> {value}', styles['Normal']))

    doc.build(elements)
    return buffer.getvalue()
